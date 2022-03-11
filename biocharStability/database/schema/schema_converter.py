from dataclasses import fields
import numpy as np
import pandas as pd
import json as j
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import re
from json2table import convert
from json2table import *

fp = 'schema.json'
toExcel = 'schema_fromJson.xlsx'
toHTML = 'schema_fromJSON.html'
toJsonVal = 'schema-val.json' # for testing
toExcelDb = 'biochar_incubation_database_init.xlsx'

def excel2json(dbfp='C:/github/biocharStability/biocharStability/database/biochar_incubation_database_2022-03-22_v0.xlsx', toJson = 'database-schema.json', tables=['articles', 'data', 'metadata', 'validation']):
    '''
    Builds a json schema of the database from the Excel database file, by reading the headers
    articles > 4 rows
    data > 5 rows
    metadata > 7 rows
    validation > 2 rows
    # common denominator : starts in A4, and then A4+x is empty
    '''
    schema = {}

    # extract table informations from excel file
    wb = load_workbook(dbfp)
    for sheet in wb.worksheets:
        #print(sheet)
        #print(sheet['B1'].value)
        if sheet.title in tables:
            schema[sheet.title] = {}
            schema[sheet.title]['description'] = sheet['B2'].value
            schema[sheet.title]['fields'] = []

            fields = []
            n = 0
            for cell in sheet['A4:A15']:
                if cell[0].value is None:
                    break
                n+=1
                fields.append(cell[0].value)
            #print(fields)
            for col in sheet.iter_cols(min_col=2, min_row=4, max_row=4+n-1, max_col=sheet.max_column):
            #     print(col)
                d = { field:cell.value for (field, cell) in zip(fields, col) }
            #     #for field, cell in zip(fields, col):
            #     #    d[field] = cell.value
                #print(d)
                schema[sheet.title]['fields'].append(d)


    with open(toJson, "w", encoding="utf-8") as s:
        j.dump(schema, s, indent=4, ensure_ascii=False)


def json2jsonval():
    '''
    Dupplicates the json schema & adds metadata validation fields in a new table
    The 2 validations fields are `fieldName_comment` and `fieldName_loci`
    Field structure: 
    { "name": "FieldName",	"type":"string" }
    Some fields are excluded: the one with group 'identification'
    '''
    with open(fp) as f:
        schema = j.load(f)

    fields = schema['metadata']['fields']
    schema['metadata_validation']['fields'] = []
    for field in fields:
        if field['group'] != "identification":
            fn1 = field['name']+'_comment'
            fn2 = field['name']+'_loci'
            schema['metadata_validation']['fields'].append({"name": fn1, "type":"string"})
            schema['metadata_validation']['fields'].append({"name": fn2, "type":"string"})

    with open(toJsonVal, "w") as o:
        j.dump(schema, o, indent=4, ensure_ascii=False) # if ensure_ascii = True, symbols like degree will be convert to their ascii code (check in Excel what's best supported)


def json2excel_schema():
    with open(toJsonVal) as f:
        schema = j.load(f)

    wb = Workbook()
    # from json-dict to xlsx
    for table, content in schema.items():
        ws = wb.create_sheet(title=table, index=None)
        # information about table
        ws['A1'] = "Table"
        ws['B1'] = table.capitalize()
        ws['A2'] = "Description"
        ws['B2'] = content['description'].capitalize()
        
        ws['A3'] = ''
        # content of table
        fields = content['fields'] # fields as list # field as dict
        df = pd.DataFrame(fields)         
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for cell in ws['A'] + ws[4]:
            cell.font = Font(bold=True)

        # column width
        dims = {}
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    if 'Sheet' in wb:
        wb.remove(wb['Sheet']) 
    wb.save(toExcel) # warning it overwrites the file fully

def json2html_schema(toJsonVal, toHTML):
    '''
    Next step:
    get table html code from dataframe df.to_html(classes='table table-stripped') 
    https://www.geeksforgeeks.org/how-to-render-pandas-dataframe-as-html-table/
    https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.to_html.html
    For nice styling and interactivity https://www.dynatable.com/#demo 
    '''
    with open(toJsonVal) as f:
        schema = j.load(f)

    json_object = schema
    build_direction = "LEFT_TO_RIGHT" #LEFT_TO_RIGHT TOP_TO_BOTTOM
    table_attributes = {
        "style" : "width:90%; border-collapse:collapse; vertical-align: top; text-align: left",
         "class" : "table table-striped "
         }

    text = '''
    <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@4.0.0/dist/css/bootstrap.min.css" integrity="sha384-Gn5384xqQ1aoWXA+058RXPxPg6fy4IWvTNh0E263XmFcJlSAwiGgFAW/dAiS6JXm" crossorigin="anonymous">

            <style>
                tr {
                    border: 1px solid black;
                    vertical-align: top; 
                    text-align: left
                }
                .container{
                    margin-left: 5%;
                    
                }
            </style>
        </head>
        <body>
            <h1>Biochar Stability Database Schema</h1>
            
            <div class="container">
            <p> The database contains 4 tables described below: articles, data, metadata, and validation.
            </p>
    '''
    for table, content in schema.items():
            print(table)
            descr = content['description']
            del content['description']
            #content['fields']
            #newdict={}
            #for k,v in [(key,d[key]) for d in content['fields'] for key in d]:
            #    if k not in newdict: newdict[k]=[v]
            #    else: newdict[k].append(v)
            subhtml = convert(content, build_direction="LEFT_TO_RIGHT", table_attributes=table_attributes)
            #r1 = '<table style="width:90%; border-collapse:collapse; vertical-align: top; text-align: left" class="table table-striped ">(.*?)<\\/table>'
            #subsub = []
            #subsub = re.findall(r1, subhtml)
            #print(subsub)
            subhtml = subhtml.removeprefix('<table style="width:90%; border-collapse:collapse; vertical-align: top; text-align: left" class="table table-striped "><tr><th>fields</th><td>')
            subhtml = subhtml.removesuffix('</td></tr></table>')
            text+= ''' <h2> '''+table+''' </h2>
                    <p> '''+descr+''' </p>
                    <p> '''+subhtml+''' </p>
                    <br />
             '''
    
    text+= '''
        </div>

        <!-- Optional JavaScript -->
        <!-- jQuery first, then Popper.js, then Bootstrap JS 
        <script src="https://code.jquery.com/jquery-3.2.1.slim.min.js" integrity="sha384-KJ3o2DKtIkvYIK3UENzmM7KCkRr/rE9/Qpg6aAZGJwFDMVNA/GpGFF93hXpG5KkN" crossorigin="anonymous"></script>
        <script src="https://cdn.jsdelivr.net/npm/popper.js@1.12.9/dist/umd/popper.min.js" integrity="sha384-ApNbgh9B+Y1QKtv3Rn7W3mgPxhU9K/ScQsAP7hUibX39j7fakFPskvXusvfa0b4Q" crossorigin="anonymous"></script>
        <script src="https://cdn.jsdelivr.net/npm/bootstrap@4.0.0/dist/js/bootstrap.min.js" integrity="sha384-JZR6Spejh4U02d8jOt6vLEHfe/JQGiRRSQQxSfFWpi1MquVdAyjUar5+76PVCmYl" crossorigin="anonymous"></script>
        -->
        </body>
    </html>
    '''
    with open(toHTML, "w") as o:
        o.write(text)
        o.close()

def json2excel_db():
    '''
    Based on the json schema or json schema validation, initialise the Excel database file where data is to be inserted.
    '''
    with open(toJsonVal) as f:
        schema = j.load(f)

    wb = Workbook()
    # from json-dict to xlsx
    for table, content in schema.items():
        ws = wb.create_sheet(title=table, index=None)
        # information about table
        ws['A1'] = "Table"
        ws['B1'] = table.capitalize()
        ws['A2'] = "Description"
        ws['B2'] = content['description']
        
        #ws['A3'] = ''

        # content of table
        fields = content['fields'] # fields as list # field as dict
        df = pd.DataFrame(fields).T 
        nrows = len(df)

        for r in dataframe_to_rows(df, index=True, header=False):
            ws.append(r)

        # bold cells
        for cell in ws['A'] + ws[4]:
            cell.font = Font(bold=True)

         # border on rows
        thin = Side(border_style="thin", color="000000")
        for i in range(nrows):
            for cell in ws[4+i]:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
       
        # column width
        dims = {}
        for cell in ws[4]:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))+5    
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    if 'Sheet' in wb:
        wb.remove(wb['Sheet']) 
    wb.save(toExcelDb) # warning it overwrites the file fully




# Run the stuff  - from json to init db and schema files
#json2jsonval()
#json2html_schema()
#json2excel_schema()
#json2excel_db()

# Run the stuff - from existing db, to json and html schema
excel2json(dbfp='C:/github/biocharStability/biocharStability/database/biochar_incubation_database_2022-03-22_live.xlsx',
           toJson = 'database-schema.json',
           tables=['articles', 'data', 'metadata', 'validation']) 
json2html_schema(toJsonVal='database-schema.json', toHTML='database-schema.html')
